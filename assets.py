"""
DIRPF annual return helpers: bank accounts, capital gains, and cryptocurrency.

These declarations are separate from Carnê-Leão (monthly salary) and go in the
annual Declaração de Imposto de Renda Pessoa Física (DIRPF):
  - Sheet "Contas_Bancarias" → Bens e Direitos (foreign bank accounts, codes 61/62)
  - Sheet "Ganhos_Capital"   → Ganhos de Capital no Exterior (stocks, ETFs, crypto disposals)
  - Sheet "Criptomoedas"     → Bens e Direitos código 89 (year-end crypto holdings at cost)

Exchange rates used (per Receita Federal guidance):
  - Bank accounts (year-end balance): BCB PTAX on 31 Dec (or last prior business day)
  - Capital gains: BCB PTAX on each transaction date (acquisition and disposal)
  - Crypto holdings: BCB PTAX on acquisition date (cost-basis reporting)
  All EUR amounts use ECB EUR/USD × BCB USD/BRL compra, same as salary conversion.
"""
import re
from datetime import date, datetime
from decimal import ROUND_HALF_EVEN, Decimal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ptax import get_spot_rates
from spreadsheet import InputError

# Styling constants — same palette as spreadsheet.py
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
_DESC_FONT   = Font(italic=True, color='595959')
_DESC_FILL   = PatternFill(fill_type='solid', fgColor='D9E1F2')
_ALT_FILL    = PatternFill(fill_type='solid', fgColor='F2F2F2')
_CENTER      = Alignment(horizontal='center')
_BRL_FMT     = 'R$ #,##0.00'
_EUR_FMT     = '#,##0.00'
_RATE_FMT    = '0.0000'
_DATE_FMT    = 'DD/MM/YYYY'
_QTY_FMT     = '#,##0.########'


def _d(value) -> Decimal:
    if pd.isna(value):
        return Decimal('0')
    return Decimal(str(value))


def _round_brl(value: Decimal) -> float:
    return float(value.quantize(Decimal('0.01'), rounding=ROUND_HALF_EVEN))


def _normalize_col(name: str) -> str:
    return re.sub(r'\s+', '_', str(name).strip())


def _parse_date(val) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s == 'nan':
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise InputError(f"Cannot parse date '{val}' — use DD/MM/YYYY format")


def _parse_eur(val) -> float:
    if val is None:
        return float('nan')
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == 'nan':
        return float('nan')
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        raise InputError(f"Cannot parse EUR value '{val}'")


# ---------------------------------------------------------------------------
# Template sheet builders
# ---------------------------------------------------------------------------

def _write_template_header(ws, headers, descriptions, col_widths):
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=i, value=desc)
        cell.font = _DESC_FONT
        cell.fill = _DESC_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 30


def add_bank_accounts_sheet(wb, year: int):
    """Add 'Contas_Bancarias' sheet: year-end EUR balances → Bens e Direitos."""
    ws = wb.create_sheet('Contas_Bancarias')
    headers = ['Banco', 'IBAN', 'Descricao', 'Saldo_EUR']
    descriptions = [
        'Nome do banco (ex: ING Belgium)',
        'Número IBAN da conta',
        'Descrição para DIRPF (ex: Conta corrente — Bélgica)',
        f'Saldo em 31/12/{year} — EUR',
    ]
    col_widths = [30, 32, 50, 20]
    _write_template_header(ws, headers, descriptions, col_widths)
    for row in range(3, 20):
        ws.cell(row=row, column=4).number_format = _EUR_FMT
        if row % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = _ALT_FILL
    ws.freeze_panes = 'A3'


def add_capital_gains_sheet(wb):
    """Add 'Ganhos_Capital' sheet: asset disposals → Ganhos de Capital no Exterior."""
    ws = wb.create_sheet('Ganhos_Capital')
    headers = ['Descricao', 'Data_Aquisicao', 'Custo_EUR', 'Data_Alienacao', 'Receita_EUR']
    descriptions = [
        'Ativo alienado (ex: Ações Apple Inc. / Bitcoin)',
        'Data de aquisição (DD/MM/AAAA)',
        'Custo total de aquisição — EUR',
        'Data de alienação/venda (DD/MM/AAAA)',
        'Receita total da alienação — EUR',
    ]
    col_widths = [50, 24, 22, 24, 22]
    _write_template_header(ws, headers, descriptions, col_widths)
    for row in range(3, 30):
        ws.cell(row=row, column=2).number_format = _DATE_FMT
        ws.cell(row=row, column=3).number_format = _EUR_FMT
        ws.cell(row=row, column=4).number_format = _DATE_FMT
        ws.cell(row=row, column=5).number_format = _EUR_FMT
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = _ALT_FILL
    ws.freeze_panes = 'A3'


def add_crypto_sheet(wb, year: int):
    """Add 'Criptomoedas' sheet: year-end holdings at cost → Bens e Direitos código 89."""
    ws = wb.create_sheet('Criptomoedas')
    headers = ['Nome', 'Ticker', 'Quantidade', 'Custo_Aquisicao_EUR', 'Data_Aquisicao']
    descriptions = [
        'Nome da criptomoeda (ex: Bitcoin)',
        'Símbolo/ticker (ex: BTC)',
        f'Quantidade detida em 31/12/{year}',
        'Custo total de aquisição — EUR (valor pago, não valor de mercado)',
        'Data média/representativa de aquisição (DD/MM/AAAA)',
    ]
    col_widths = [26, 14, 20, 34, 28]
    _write_template_header(ws, headers, descriptions, col_widths)
    for row in range(3, 20):
        ws.cell(row=row, column=3).number_format = _QTY_FMT
        ws.cell(row=row, column=4).number_format = _EUR_FMT
        ws.cell(row=row, column=5).number_format = _DATE_FMT
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = _ALT_FILL
    ws.freeze_panes = 'A3'


def add_asset_sheets_to_template(path: str, year: int):
    """Open an existing workbook and append all three asset declaration sheets."""
    wb = load_workbook(path)
    add_bank_accounts_sheet(wb, year)
    add_capital_gains_sheet(wb)
    add_crypto_sheet(wb, year)
    wb.save(path)


# ---------------------------------------------------------------------------
# Read helpers
# ---------------------------------------------------------------------------

def read_bank_accounts(path: str):
    """Read 'Contas_Bancarias' sheet. Returns DataFrame or None if absent/empty."""
    try:
        df = pd.read_excel(path, sheet_name='Contas_Bancarias', header=0, skiprows=[1])
    except Exception:
        return None
    df.columns = [_normalize_col(c) for c in df.columns]
    df = df.dropna(how='all')
    if df.empty or 'Saldo_EUR' not in df.columns:
        return None
    df = df.dropna(subset=['Saldo_EUR'])
    if df.empty:
        return None
    df['Saldo_EUR'] = df['Saldo_EUR'].apply(_parse_eur)
    return df.reset_index(drop=True)


def read_capital_gains(path: str):
    """Read 'Ganhos_Capital' sheet. Returns only rows where all five fields are present."""
    try:
        df = pd.read_excel(path, sheet_name='Ganhos_Capital', header=0, skiprows=[1])
    except Exception:
        return None
    df.columns = [_normalize_col(c) for c in df.columns]
    df = df.dropna(how='all')
    required = ['Descricao', 'Data_Aquisicao', 'Custo_EUR', 'Data_Alienacao', 'Receita_EUR']
    if any(c not in df.columns for c in required):
        return None
    df = df.dropna(subset=required)
    if df.empty:
        return None
    df['Data_Aquisicao'] = df['Data_Aquisicao'].apply(_parse_date)
    df['Data_Alienacao'] = df['Data_Alienacao'].apply(_parse_date)
    df['Custo_EUR']      = df['Custo_EUR'].apply(_parse_eur)
    df['Receita_EUR']    = df['Receita_EUR'].apply(_parse_eur)
    return df.reset_index(drop=True)


def read_crypto(path: str):
    """Read 'Criptomoedas' sheet. Returns DataFrame or None if absent/empty."""
    try:
        df = pd.read_excel(path, sheet_name='Criptomoedas', header=0, skiprows=[1])
    except Exception:
        return None
    df.columns = [_normalize_col(c) for c in df.columns]
    df = df.dropna(how='all')
    required = ['Nome', 'Custo_Aquisicao_EUR', 'Data_Aquisicao']
    if any(c not in df.columns for c in required):
        return None
    df = df.dropna(subset=required)
    if df.empty:
        return None
    df['Custo_Aquisicao_EUR'] = df['Custo_Aquisicao_EUR'].apply(_parse_eur)
    df['Data_Aquisicao']      = df['Data_Aquisicao'].apply(_parse_date)
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Conversion
# ---------------------------------------------------------------------------

def convert_bank_accounts(df: pd.DataFrame, year: int, cache_path: str) -> pd.DataFrame:
    """Convert EUR year-end balances to BRL using 31 Dec spot rates."""
    target = date(year, 12, 31)
    r = get_spot_rates(target, cache_path)
    fx = _d(r['ecb_eur_usd']) * _d(r['bcb_usd_brl'])
    df = df.copy()
    df['Saldo_BRL']   = df['Saldo_EUR'].apply(lambda v: _round_brl(_d(v) * fx))
    df['ECB_EUR_USD'] = r['ecb_eur_usd']
    df['ECB_Data']    = r['ecb_date']
    df['BCB_USD_BRL'] = r['bcb_usd_brl']
    df['BCB_Data']    = r['bcb_date']
    df['Observacoes'] = r['notes']
    return df


def convert_capital_gains(df: pd.DataFrame, cache_path: str) -> pd.DataFrame:
    """
    Fetch spot rates for each acquisition and disposal date, then compute BRL gain/loss.
    Rate rule: BCB PTAX on the transaction date (with up-to-7-day fallback).
    """
    df = df.copy()
    custo_brl, receita_brl, ganho_brl = [], [], []
    ecb_acq, bcb_acq, ecb_ali, bcb_ali, notes_list = [], [], [], [], []

    for _, row in df.iterrows():
        r_acq = get_spot_rates(row['Data_Aquisicao'], cache_path)
        r_ali = get_spot_rates(row['Data_Alienacao'], cache_path)

        fx_acq = _d(r_acq['ecb_eur_usd']) * _d(r_acq['bcb_usd_brl'])
        fx_ali = _d(r_ali['ecb_eur_usd']) * _d(r_ali['bcb_usd_brl'])

        c_brl = _round_brl(_d(row['Custo_EUR'])   * fx_acq)
        r_brl = _round_brl(_d(row['Receita_EUR']) * fx_ali)
        g_brl = _round_brl(Decimal(str(r_brl)) - Decimal(str(c_brl)))

        custo_brl.append(c_brl)
        receita_brl.append(r_brl)
        ganho_brl.append(g_brl)
        ecb_acq.append(r_acq['ecb_eur_usd'])
        bcb_acq.append(r_acq['bcb_usd_brl'])
        ecb_ali.append(r_ali['ecb_eur_usd'])
        bcb_ali.append(r_ali['bcb_usd_brl'])
        parts = [n for n in (r_acq['notes'], r_ali['notes']) if n != 'OK']
        notes_list.append('; '.join(parts) if parts else 'OK')

    df['Custo_BRL']     = custo_brl
    df['Receita_BRL']   = receita_brl
    df['Ganho_BRL']     = ganho_brl
    df['ECB_Aquisicao'] = ecb_acq
    df['BCB_Aquisicao'] = bcb_acq
    df['ECB_Alienacao'] = ecb_ali
    df['BCB_Alienacao'] = bcb_ali
    df['Observacoes']   = notes_list
    return df


def convert_crypto(df: pd.DataFrame, cache_path: str) -> pd.DataFrame:
    """
    Convert crypto acquisition costs to BRL using acquisition date spot rates.
    Brazilian rule: declare at acquisition cost in BRL, not current market value.
    """
    df = df.copy()
    custo_brl, ecb_list, bcb_list, notes_list = [], [], [], []

    for _, row in df.iterrows():
        r = get_spot_rates(row['Data_Aquisicao'], cache_path)
        fx = _d(r['ecb_eur_usd']) * _d(r['bcb_usd_brl'])
        custo_brl.append(_round_brl(_d(row['Custo_Aquisicao_EUR']) * fx))
        ecb_list.append(r['ecb_eur_usd'])
        bcb_list.append(r['bcb_usd_brl'])
        notes_list.append(r['notes'])

    df['Custo_BRL']   = custo_brl
    df['ECB_EUR_USD'] = ecb_list
    df['BCB_USD_BRL'] = bcb_list
    df['Observacoes'] = notes_list
    return df


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

def _write_output_header(ws, headers, col_widths):
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_output_rows(ws, df, data_cols, col_fmts, start_row=2):
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (dcol, fmt) in enumerate(zip(data_cols, col_fmts), 1):
            val = row.get(dcol)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fmt and val is not None:
                cell.number_format = fmt
            if fill:
                cell.fill = fill


def write_assets_output(
    output_path: str,
    year: int,
    df_banks=None,
    df_gains=None,
    df_crypto=None,
):
    """Append asset declaration sheets to the existing output workbook."""
    wb = load_workbook(output_path)
    if df_banks is not None:
        _write_bank_accounts_sheet(wb, df_banks, year)
    if df_gains is not None:
        _write_capital_gains_sheet(wb, df_gains, year)
    if df_crypto is not None:
        _write_crypto_sheet(wb, df_crypto, year)
    wb.save(output_path)


def _write_bank_accounts_sheet(wb, df: pd.DataFrame, year: int):
    ws = wb.create_sheet(f'Bens_Direitos_{year}')
    headers   = ['Banco', 'IBAN', 'Descricao', 'Saldo_EUR', 'Saldo_BRL',
                 'ECB_EUR_USD', 'ECB_Data', 'BCB_USD_BRL', 'BCB_Data', 'Observacoes']
    col_widths = [30, 32, 50, 18, 18, 13, 13, 13, 13, 42]
    col_fmts   = [None, None, None, _EUR_FMT, _BRL_FMT,
                  _RATE_FMT, _DATE_FMT, _RATE_FMT, _DATE_FMT, None]
    _write_output_header(ws, headers, col_widths)
    _write_output_rows(ws, df, headers, col_fmts)
    ws.freeze_panes = 'A2'


def _write_capital_gains_sheet(wb, df: pd.DataFrame, year: int):
    ws = wb.create_sheet(f'Ganhos_Capital_{year}')
    headers    = ['Descricao', 'Data_Aquisicao', 'Custo_EUR', 'Custo_BRL',
                  'Data_Alienacao', 'Receita_EUR', 'Receita_BRL', 'Ganho_BRL',
                  'ECB_Aquisicao', 'BCB_Aquisicao', 'ECB_Alienacao', 'BCB_Alienacao',
                  'Observacoes']
    col_widths = [50, 18, 16, 16, 18, 16, 16, 16, 14, 14, 14, 14, 42]
    col_fmts   = [None, _DATE_FMT, _EUR_FMT, _BRL_FMT,
                  _DATE_FMT, _EUR_FMT, _BRL_FMT, _BRL_FMT,
                  _RATE_FMT, _RATE_FMT, _RATE_FMT, _RATE_FMT, None]
    _write_output_header(ws, headers, col_widths)
    _write_output_rows(ws, df, headers, col_fmts)
    # Total row
    total_row = len(df) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        if h in ('Custo_BRL', 'Receita_BRL', 'Ganho_BRL'):
            cell = ws.cell(row=total_row, column=col_idx, value=df[h].sum())
            cell.number_format = _BRL_FMT
            cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def _write_crypto_sheet(wb, df: pd.DataFrame, year: int):
    ws = wb.create_sheet(f'Criptomoedas_{year}')
    headers    = ['Nome', 'Ticker', 'Quantidade', 'Custo_Aquisicao_EUR', 'Data_Aquisicao',
                  'Custo_BRL', 'ECB_EUR_USD', 'BCB_USD_BRL', 'Observacoes']
    col_widths = [26, 14, 20, 26, 22, 18, 13, 13, 42]
    col_fmts   = [None, None, _QTY_FMT, _EUR_FMT, _DATE_FMT,
                  _BRL_FMT, _RATE_FMT, _RATE_FMT, None]
    _write_output_header(ws, headers, col_widths)
    _write_output_rows(ws, df, headers, col_fmts)
    ws.freeze_panes = 'A2'
