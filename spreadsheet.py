"""Input template creation, input reading, and output writing."""
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class InputError(Exception):
    pass


EXPECTED_COLUMNS = [
    'Mes',
    'Data_Pagamento',
    'Salario_Bruto_EUR',
    'Previdencia_Social_EUR',
    'Imposto_Retido_Belgica_EUR',
    'Opcoes_Acoes_EUR',
    'Imposto_Retido_Opcoes_EUR',
    'Vakantiegeld_EUR',
    'Imposto_Retido_Vakantiegeld_EUR',
    'Bonus_13e_Maand_EUR',
    'Previdencia_Social_13e_Maand_EUR',
    'Imposto_Retido_13e_Maand_EUR',
    'Salario_Liquido_EUR',
]

# Columns that are optional (absent in old templates → filled with NaN)
OPTIONAL_COLUMNS = {
    'Opcoes_Acoes_EUR',
    'Imposto_Retido_Opcoes_EUR',
    'Vakantiegeld_EUR',
    'Imposto_Retido_Vakantiegeld_EUR',
    'Bonus_13e_Maand_EUR',
    'Previdencia_Social_13e_Maand_EUR',
    'Imposto_Retido_13e_Maand_EUR',
}

_HEADER_FONT  = Font(bold=True, color='FFFFFF')
_HEADER_FILL  = PatternFill(fill_type='solid', fgColor='1F4E79')
_DESC_FONT    = Font(italic=True, color='595959')
_DESC_FILL    = PatternFill(fill_type='solid', fgColor='D9E1F2')
_SUBHEAD_FONT = Font(bold=True, color='1F4E79')
_ALT_FILL     = PatternFill(fill_type='solid', fgColor='F2F2F2')
_CENTER       = Alignment(horizontal='center')
_BRL_FMT      = 'R$ #,##0.00'
_EUR_FMT      = '#,##0.00'
_RATE_FMT     = '0.0000'
_DATE_FMT     = 'DD/MM/YYYY'

_BORDER_DARK  = '1F4E79'
_BORDER_LIGHT = 'B4C7E7'
_THICK_SIDE   = Side(border_style='medium', color=_BORDER_DARK)
_THIN_SIDE    = Side(border_style='thin',   color=_BORDER_LIGHT)


def _apply_group_box(ws, r1: int, r2: int, c1: int, c2: int, header_sep_row: int = 2):
    """Draw a thick outer border, thin inner verticals, and a thin band around the sub-header row."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            top    = _THICK_SIDE if r == r1 else (_THIN_SIDE if r == header_sep_row else None)
            bottom = _THICK_SIDE if r == r2 else (_THIN_SIDE if r == header_sep_row else None)
            left   = _THICK_SIDE if c == c1 else _THIN_SIDE
            right  = _THICK_SIDE if c == c2 else _THIN_SIDE
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


# ---------------------------------------------------------------------------
# Template creation
# ---------------------------------------------------------------------------

def create_template(path: str, year: int):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Salarios_EUR'

    headers = [
        'Mes',
        'Data_Pagamento',
        'Salario_Bruto_EUR',
        'Previdencia_Social_EUR',
        'Imposto_Retido_Belgica_EUR',
        'Opcoes_Acoes_EUR',
        'Imposto_Retido_Opcoes_EUR',
        'Vakantiegeld_EUR',
        'Imposto_Retido_Vakantiegeld_EUR',
        'Bonus_13e_Maand_EUR',
        'Previdencia_Social_13e_Maand_EUR',
        'Imposto_Retido_13e_Maand_EUR',
        'Salario_Liquido_EUR',
    ]
    descriptions = [
        'Mes (1-12)',
        'Data do pagamento (DD/MM/AAAA)',
        'Salario bruto (bruto loon) — EUR',
        'Previdencia social salario (RSZ/ONSS) — EUR',
        'Imposto retido salario (bedrijfsvoorheffing) — EUR',
        'Stock options bruto — EUR (em branco se nao houver)',
        'Imposto retido stock options — EUR (em branco se nao houver)',
        'Vakantiegeld bruto — EUR (em branco se nao houver)',
        'Imposto retido vakantiegeld — EUR (em branco se nao houver)',
        '13e maand bruto — EUR (em branco se nao houver)',
        'Previdencia social 13e maand (RSZ/ONSS) — EUR (em branco se nao houver)',
        'Imposto retido 13e maand — EUR (em branco se nao houver)',
        'Salario liquido combinado (netto loon, todos os rendimentos) — EUR (verificacao)',
    ]
    col_widths = [8, 24, 28, 32, 38, 26, 34, 26, 34, 26, 36, 34, 42]

    # Header row
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(i)].width = w

    # Description row
    for i, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=i, value=desc)
        cell.font = _DESC_FONT
        cell.fill = _DESC_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Data rows: pre-fill month numbers
    n_cols = len(headers)
    for month in range(1, 13):
        row = month + 2
        ws.cell(row=row, column=1, value=month).alignment = _CENTER
        ws.cell(row=row, column=2).number_format = _DATE_FMT
        for col in range(3, n_cols + 1):
            ws.cell(row=row, column=col).number_format = _EUR_FMT
        if month % 2 == 0:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = _ALT_FILL

    ws.freeze_panes = 'B3'
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Read input
# ---------------------------------------------------------------------------

def _normalize_col(name: str) -> str:
    return re.sub(r'\s+', '_', str(name).strip())


def _parse_date(val) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s == 'nan':
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise InputError(f"Cannot parse date '{val}' — use DD/MM/YYYY format")


def _parse_eur(val) -> float:
    if val is None:
        return float('nan')
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == 'nan':
        return float('nan')
    # Handle European comma-decimal notation: "1.234,56" → 1234.56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        raise InputError(f"Cannot parse EUR value '{val}'")


def read_input(path: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name='Salarios_EUR', header=0, skiprows=[1])
    except Exception as e:
        raise InputError(f'Cannot read {path}: {e}')

    df.columns = [_normalize_col(c) for c in df.columns]

    # Required columns
    required = [c for c in EXPECTED_COLUMNS if c not in OPTIONAL_COLUMNS]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise InputError(
            f'Missing columns: {missing}\nExpected: {EXPECTED_COLUMNS}'
        )

    # Optional columns — add as NaN if not present (old template)
    for col in OPTIONAL_COLUMNS:
        if col not in df.columns:
            df[col] = float('nan')

    df = df[EXPECTED_COLUMNS].copy()

    # Drop non-data rows (Total rows, notes, etc.) — keep only valid month integers 1-12
    df['Mes'] = pd.to_numeric(df['Mes'], errors='coerce')
    df = df[df['Mes'].between(1, 12)].copy()

    # Check duplicate months
    dupes = df['Mes'][df['Mes'].duplicated()].astype(int).tolist()
    if dupes:
        raise InputError(f'Duplicate entries for month(s): {dupes}')

    # Parse dates
    df['Data_Pagamento'] = df['Data_Pagamento'].apply(_parse_date)

    # Parse EUR amounts
    for col in ['Salario_Bruto_EUR', 'Previdencia_Social_EUR',
                'Imposto_Retido_Belgica_EUR', 'Opcoes_Acoes_EUR',
                'Imposto_Retido_Opcoes_EUR', 'Vakantiegeld_EUR',
                'Imposto_Retido_Vakantiegeld_EUR', 'Bonus_13e_Maand_EUR',
                'Previdencia_Social_13e_Maand_EUR', 'Imposto_Retido_13e_Maand_EUR',
                'Salario_Liquido_EUR']:
        df[col] = df[col].apply(_parse_eur)

    return df


# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------

def write_output(df: pd.DataFrame, path: str, year: int):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Carne_Leao_{year}'

    # Per-item groups: each income type gets 5 columns boxed together.
    item_sub = ['Rendimento', 'Prev.Social', 'Tributavel', 'Imp.Retido', 'Netto']
    item_w   = [14, 14, 14, 14, 14]
    item_fmt = [_BRL_FMT] * 5

    groups = [
        ('Salario', item_sub, item_w,
            ['rendimentos_brl', 'deducao_prev_brl', 'tributavel_salario_brl',
             'imposto_retido_brl', 'netto_salario_brl'],
            item_fmt),
        ('Opcoes', item_sub, item_w,
            ['rendimentos_opcoes_brl', None, 'rendimentos_opcoes_brl',
             'imposto_opcoes_brl', 'netto_opcoes_brl'],
            item_fmt),
        ('Vakantiegeld', item_sub, item_w,
            ['rendimentos_vakantiegeld_brl', None, 'rendimentos_vakantiegeld_brl',
             'imposto_vakantiegeld_brl', 'netto_vakantiegeld_brl'],
            item_fmt),
        ('13e Maand', item_sub, item_w,
            ['rendimentos_13e_maand_brl', 'deducao_prev_13e_maand_brl',
             'tributavel_13e_maand_brl', 'imposto_13e_maand_brl',
             'netto_13e_maand_brl'],
            item_fmt),
        ('Cambio', ['ECB EUR/USD', 'BCB USD/BRL', 'ECB Data', 'BCB Data'],
            [13, 13, 12, 12],
            ['ecb_eur_usd', 'bcb_usd_brl', 'ecb_date', 'bcb_date'],
            [_RATE_FMT, _RATE_FMT, _DATE_FMT, _DATE_FMT]),
        ('Resumo', ['Liquido Input', 'Base Calculo', 'Observacoes'],
            [16, 16, 30],
            ['salario_liquido_brl', 'base_calculo_brl', 'notes'],
            [_BRL_FMT, _BRL_FMT, None]),
    ]

    # Mes column — merged A1:A2
    for r in (1, 2):
        c = ws.cell(row=r, column=1)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    ws.cell(row=1, column=1, value='Mes')
    ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    ws.column_dimensions[get_column_letter(1)].width = 6

    # Group headers (row 1) + sub-headers (row 2)
    col = 2
    group_ranges = []
    all_keys = ['Mes']
    all_fmts = [None]

    for label, sub_headers, widths, data_keys, fmts in groups:
        c1 = col
        c2 = col + len(sub_headers) - 1
        # Row 1 — group label (merged across the group's columns, with fill on every constituent cell)
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
        ws.cell(row=1, column=c1, value=label)
        if c2 > c1:
            ws.merge_cells(start_row=1, end_row=1, start_column=c1, end_column=c2)
        # Row 2 — short sub-headers + column widths
        for i, (h, w) in enumerate(zip(sub_headers, widths)):
            sc = ws.cell(row=2, column=c1 + i, value=h)
            sc.font = _SUBHEAD_FONT
            sc.fill = _DESC_FILL
            sc.alignment = _CENTER
            ws.column_dimensions[get_column_letter(c1 + i)].width = w
        group_ranges.append((c1, c2))
        all_keys.extend(data_keys)
        all_fmts.extend(fmts)
        col = c2 + 1

    # Data rows (start at row 3)
    n_rows = len(df)
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (dkey, fmt) in enumerate(zip(all_keys, all_fmts), 1):
            val = row.get(dkey) if dkey is not None else None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fmt:
                cell.number_format = fmt
            if fill:
                cell.fill = fill
        # Center the Mes value
        ws.cell(row=row_idx, column=1).alignment = _CENTER

    # Borders: a box per group, plus one around Mes
    last_row = 2 + n_rows
    _apply_group_box(ws, 1, last_row, 1, 1)
    for c1, c2 in group_ranges:
        _apply_group_box(ws, 1, last_row, c1, c2)

    ws.freeze_panes = 'B3'

    # Summary sheet
    ws2 = wb.create_sheet('Resumo_Anual')
    ws2['A1'] = f'Resumo Anual {year}'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 42
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22

    _EUR_FMT_SUM = '[$EUR] #,##0.00'

    # Column headers
    ws2.cell(row=2, column=2, value='EUR').font = Font(bold=True)
    ws2.cell(row=2, column=3, value='BRL').font = Font(bold=True)

    summary_rows = [
        (
            'Total Rendimentos Salario',
            df['Salario_Bruto_EUR'].sum(),
            df['rendimentos_brl'].sum(),
        ),
        (
            'Total Rendimentos Stock Options',
            df['Opcoes_Acoes_EUR'].sum(),
            df['rendimentos_opcoes_brl'].sum(),
        ),
        (
            'Total Rendimentos Vakantiegeld',
            df['Vakantiegeld_EUR'].sum(),
            df['rendimentos_vakantiegeld_brl'].sum(),
        ),
        (
            'Total Rendimentos 13e Maand',
            df['Bonus_13e_Maand_EUR'].sum(),
            df['rendimentos_13e_maand_brl'].sum(),
        ),
        (
            'Total Deducoes Prev. Social Salario',
            df['Previdencia_Social_EUR'].sum(),
            df['deducao_prev_brl'].sum(),
        ),
        (
            'Total Deducoes Prev. Social 13e Maand',
            df['Previdencia_Social_13e_Maand_EUR'].sum(),
            df['deducao_prev_13e_maand_brl'].sum(),
        ),
        (
            'Total Imposto Retido Salario',
            df['Imposto_Retido_Belgica_EUR'].sum(),
            df['imposto_retido_brl'].sum(),
        ),
        (
            'Total Imposto Retido Stock Options',
            df['Imposto_Retido_Opcoes_EUR'].sum(),
            df['imposto_opcoes_brl'].sum(),
        ),
        (
            'Total Imposto Retido Vakantiegeld',
            df['Imposto_Retido_Vakantiegeld_EUR'].sum(),
            df['imposto_vakantiegeld_brl'].sum(),
        ),
        (
            'Total Imposto Retido 13e Maand',
            df['Imposto_Retido_13e_Maand_EUR'].sum(),
            df['imposto_13e_maand_brl'].sum(),
        ),
        (
            'Total Base de Calculo',
            (df['Salario_Bruto_EUR'].fillna(0) + df['Opcoes_Acoes_EUR'].fillna(0)
             + df['Vakantiegeld_EUR'].fillna(0) + df['Bonus_13e_Maand_EUR'].fillna(0)
             - df['Previdencia_Social_EUR'].fillna(0)
             - df['Previdencia_Social_13e_Maand_EUR'].fillna(0)).sum(),
            df['base_calculo_brl'].sum(),
        ),
    ]
    for i, (label, eur_val, brl_val) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        eur_cell = ws2.cell(row=i, column=2, value=eur_val)
        eur_cell.number_format = _EUR_FMT_SUM
        brl_cell = ws2.cell(row=i, column=3, value=brl_val)
        brl_cell.number_format = _BRL_FMT

    # Blank row, then combined taxable income and total taxes paid
    extra_start = 3 + len(summary_rows) + 1
    extra_rows = [
        (
            'Total Tributavel (Salario + Opcoes + Vakantiegeld + 13e Maand)',
            (df['Salario_Bruto_EUR'].fillna(0) + df['Opcoes_Acoes_EUR'].fillna(0)
             + df['Vakantiegeld_EUR'].fillna(0) + df['Bonus_13e_Maand_EUR'].fillna(0)
             - df['Previdencia_Social_EUR'].fillna(0)
             - df['Previdencia_Social_13e_Maand_EUR'].fillna(0)).sum(),
            (df['tributavel_salario_brl'] + df['rendimentos_opcoes_brl']
             + df['rendimentos_vakantiegeld_brl']
             + df['tributavel_13e_maand_brl']).sum(),
        ),
        (
            'Total Imposto Retido (Todos)',
            (df['Imposto_Retido_Belgica_EUR'].fillna(0)
             + df['Imposto_Retido_Opcoes_EUR'].fillna(0)
             + df['Imposto_Retido_Vakantiegeld_EUR'].fillna(0)
             + df['Imposto_Retido_13e_Maand_EUR'].fillna(0)).sum(),
            (df['imposto_retido_brl'] + df['imposto_opcoes_brl']
             + df['imposto_vakantiegeld_brl'] + df['imposto_13e_maand_brl']).sum(),
        ),
    ]
    for i, (label, eur_val, brl_val) in enumerate(extra_rows, extra_start):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        eur_cell = ws2.cell(row=i, column=2, value=eur_val)
        eur_cell.number_format = _EUR_FMT_SUM
        brl_cell = ws2.cell(row=i, column=3, value=brl_val)
        brl_cell.number_format = _BRL_FMT

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
